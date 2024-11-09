from tkinter.font import Font
from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Count, Sum
from django.utils import timezone
from django.db.models.functions import ExtractHour
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView
from shop.models import Order
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
class StatisticsView(LoginRequiredMixin, TemplateView):

    template_name = 'manager/statistics/statistics.html'
    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)


class OrderStatisticsByDateView(APIView):
    def get(self, request, *args, **kwargs):
        date_str = request.GET.get('date', timezone.now().date().strftime('%Y-%m-%d'))
        date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        orders_on_date = Order.objects.filter(order_date__date=date)

        orders_by_hour = (
            orders_on_date
            .annotate(hour=ExtractHour('order_date'))  # Extract hour from order_date
            .values('hour')  # Group by hour
            .annotate(total_orders=Count('id'), total_sales=Sum('total_price'))
            .order_by('hour')  # Order by hour
        )

        # Prepare data for the chart
        hours = []
        total_orders = []
        total_sales = []
        for order in orders_by_hour:
            hours.append(order['hour'])
            total_orders.append(order['total_orders'])
            total_sales.append(order['total_sales'] or 0)  # Ensure sales are not None

        statistics = {
            'date': date,
            'hours': hours,  # x-axis labels for hours
            'total_orders': total_orders,  # y-axis data for total orders
            'total_sales': total_sales  # y-axis data for total sales
        }

        return Response(statistics, status=status.HTTP_200_OK)

class OrderStatisticsFromToView(APIView):
    """
    API View to return statistics based on orders for a date range (from start date to end date)
    """
    def get(self, request, *args, **kwargs):
        start_date_str = request.GET.get('start_date', timezone.now().date().strftime('%Y-%m-%d'))
        end_date_str = request.GET.get('end_date', timezone.now().date().strftime('%Y-%m-%d'))

        try:
            start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            if start_date > end_date:
                return Response({"error": "start_date cannot be after end_date."}, status=status.HTTP_400_BAD_REQUEST)

        except ValueError as e:
            return Response({"error": f"Invalid date format: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        orders_in_range = Order.objects.filter(order_date__date__gte=start_date, order_date__date__lte=end_date)
        orders_by_day = (
            orders_in_range
            .values('order_date__date')  # Group by date
            .annotate(total_orders=Count('id'), total_sales=Sum('total_price'))
            .order_by('order_date__date')  # Sort by date
        )

        # Prepare data for the chart
        dates = []
        total_orders = []
        total_sales = []
        for order in orders_by_day:
            dates.append(order['order_date__date'])
            total_orders.append(order['total_orders'])
            total_sales.append(order['total_sales'] or 0)  # Ensure sales are not None

        statistics = {
            'start_date': start_date,
            'end_date': end_date,
            'dates': dates,  # x-axis labels for dates
            'total_orders': total_orders,  # y-axis data for total orders
            'total_sales': total_sales  # y-axis data for total sales
        }

        return Response(statistics, status=status.HTTP_200_OK)

class OrderStatisticsByStatusView(APIView):
    def get(self, request, *args, **kwargs):
        orders_by_status = (
            Order.objects
            .values('status') 
            .annotate(total_orders=Count('id'), total_sales=Sum('total_price'))  
            .order_by('status')  
        )
        statistics = {
            'orders_by_status': orders_by_status
        }

        return Response(statistics, status=status.HTTP_200_OK)
class ExportOrderStatisticsToExcelView(APIView):
    """
    API View to export order statistics to an Excel file, including detailed OrderItem information.
    """

    def get(self, request, *args, **kwargs):
        # Get orders in the specified date range
        start_date_str = request.GET.get('start_date', timezone.now().date().strftime('%Y-%m-%d'))
        end_date_str = request.GET.get('end_date', timezone.now().date().strftime('%Y-%m-%d'))

        try:
            start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            if start_date > end_date:
                return Response({"error": "start_date cannot be after end_date."}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError as e:
            return Response({"error": f"Invalid date format: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        # Filter orders within date range
        orders = Order.objects.filter(order_date__date__gte=start_date, order_date__date__lte=end_date)

        # Create a workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Order Item Statistics"

        # Define headers
        headers = ["STT", "Tên Bánh", "Loại", "Ngày", "Giá", "Số Lượng", "Total"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Populate rows with order item data
        total_all = 0
        row_num = 2  # Start data rows at row 2
        for order in orders:
            for item in order.items.all():
                worksheet.cell(row=row_num, column=1, value=row_num - 1)
                worksheet.cell(row=row_num, column=2, value=item.product.name)
                worksheet.cell(row=row_num, column=3, value=item.product.category)
                worksheet.cell(row=row_num, column=4, value=order.order_date.strftime("%Y-%m-%d %H:%M:%S"))
                worksheet.cell(row=row_num, column=5, value=item.price)
                worksheet.cell(row=row_num, column=6, value=item.quantity)
                worksheet.cell(row=row_num, column=7, value=item.total_price)

                # Accumulate total
                total_all += item.total_price
                row_num += 1

        # Add Total All row at the end
        total_label_cell = worksheet.cell(row=row_num, column=6, value="Total All")
        total_label_cell.font = Font(bold=True)
        total_value_cell = worksheet.cell(row=row_num, column=7, value=total_all)
        total_value_cell.font = Font(bold=True)

        # Set column widths for better readability
        column_widths = [15, 25, 15, 20, 10, 10, 15]  # Adjust widths as needed
        for i, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

        # Save to response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="Order_Item_Statistics_{start_date}_{end_date}.xlsx"'
        workbook.save(response)
        return response
